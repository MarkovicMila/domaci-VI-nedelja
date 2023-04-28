import openpyxl as op
import pandas as pd
import psycopg2 as pg

wb=op.load_workbook('sampledatafoodinfo.xlsx')
ws=wb['FoodList']
fooditem=ws['B']
calories=ws['D']
protein=ws['E']
fat=ws['F']
carbs=ws['G']
fibre=ws['H']

s_fooditem=pd.Series([i.value for i in fooditem[1:1095]])
s_fooditem.fillna('',inplace=True)

s_calories=pd.Series([i.value for i in calories[1:1095]])
s_calories.fillna(0,inplace=True)

s_protein=pd.Series([i.value for i in protein[1:1095]])
s_protein.fillna(0,inplace=True)

s_fat=pd.Series([i.value for i in fat[1:1095]])
s_fat.fillna(0,inplace=True)

s_carbs=pd.Series([i.value for i in carbs[1:1095]])
s_carbs.fillna(0,inplace=True)

s_fibre=pd.Series([i.value for i in fibre[1:1095]])
s_fibre.fillna(0,inplace=True)

calories_avg=s_calories.mean()
protein_avg=s_protein.mean()
fat_avg=s_fat.mean()
carbs_avg=s_carbs.mean()

dt=pd.DataFrame({'Calories':[calories_avg],
                 'Protein':[protein_avg],
                 'Fat':[fat_avg],
                 'Carbs':[carbs_avg]})
dt.to_csv('Foodinfo.csv',index=False)
print(dt)


def create_table():
    try:
        con=pg.connect(database='fooddata',
                       user='postgres',
                       port='5432',
                       host='localhost',
                       password='itoip')
        cursor=con.cursor()
        cursor.execute('''
        CREATE TABLE FOOD (ID SERIAL PRIMARY KEY,
                FOODITEM VARCHAR(150),
				 CALORIES INTEGER,
				 PROTEIN INTEGER,
				 FAT INTEGER,
				 CARBS INTEGER,
				 FIBRE FLOAT);
				 ''')
        
        print('Table created successfully!')
        con.commit()

    except(Exception,pg.Error) as e:
        print('Error: ',e)
    finally:
        con.close()
        cursor.close()

def inserti_into():
    try:
        con=pg.connect(database='fooddata',
                       user='postgres',
                       port='5432',
                       host='localhost',
                       password='itoip')
        cursor=con.cursor()

        for i in range(len(s_fooditem)):
            com='''INSERT INTO FOOD (FOODITEM,CALORIES,PROTEIN,FAT,CARBS,FIBRE) VALUES ('{}',{},{},{},{},{})
            '''.format(s_fooditem[i],s_calories[i],s_protein[i],s_fat[i],s_carbs[i],s_fibre[i])
            cursor.execute(com)
        con.commit()
    except(Exception,pg.Error) as e:
        print('Error: ',e)
    finally:
        cursor.close()
        con.close()

create_table()
inserti_into()